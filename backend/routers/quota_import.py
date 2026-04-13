"""
定额 Excel 导入路由
POST /admin/quota/import — 上传 Excel 并导入定额数据
GET /admin/quota/report — 获取最新导入报告
"""
import json
import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, func
from collections import defaultdict
import openpyxl

from database import get_db
from core.security import get_current_user, require_role
from models.quota import Quota, Material

router = APIRouter(prefix="/admin/quota", tags=["定额管理-管理员"])


# ===== 内存报告存储（生产环境应换用数据库/Redis）=====
_import_report: dict = {}


# ===== 辅助函数 =====
def num(v) -> Optional[float]:
    """安全转浮点数"""
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_excel_to_records(excel_bytes: bytes) -> list[dict]:
    """将 Excel 文件解析为字典列表"""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if not headers or headers[0] is None:
        raise ValueError("Excel 文件缺少表头或表头为空")

    # 标准化表头
    std_headers = []
    for h in headers:
        if h is None:
            std_headers.append(None)
        else:
            # 去掉空白，统一大小写
            h_str = str(h).strip()
            # 映射常见别名
            alias_map = {
                "定额编号": "定额编号",
                "编号": "定额编号",
                "子目编号": "定额编号",
                "定额名称": "名称",
                "名称": "名称",
                "子目名称": "名称",
                "项目名称": "项目名称",
                "计量单位": "计量单位",
                "单位": "计量单位",
                "工作内容": "工作内容",
                "分部工程": "分部工程",
                "分部": "分部工程",
                "全费用": "全费用",
                "综合单价": "全费用",
                "基价": "全费用",
                "其中人工费": "人工费",
                "人工费": "人工费",
                "人工": "人工费",
                "材料费": "材料费",
                "机械费": "机械费",
                "机械使用费": "机械费",
                "管理费": "管理费",
                "规费含利润": "管理费",
                "增值税": "增值税",
                "税金": "增值税",
                "主材费": "主材费",
                "设备费": "设备费",
            }
            std_headers.append(alias_map.get(h_str, h_str))

    records = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        record = {}
        for header, value in zip(std_headers, row):
            if header:
                record[header] = value
        if record.get("定额编号"):
            records.append(record)

    return records


def guess_section(record: dict) -> str:
    """从记录中推断分部路径"""
    if record.get("分部工程"):
        return record["分部工程"]
    name = str(record.get("名称", "") or "")
    # 简单规则
    for kw in ["绿化", "园路", "景观", "亭", "廊", "花架", "水池"]:
        if kw in name:
            return f"园林景观 / {kw}"
    return "未分类"


def import_records(db: Session, records: list[dict]) -> dict:
    """将解析后的记录写入数据库"""
    total = len(records)
    success = 0
    skipped = 0
    errors = []
    section_dist: dict[str, int] = defaultdict(int)

    for i, r in enumerate(records):
        quota_id = str(r.get("定额编号", "")).strip()
        if not quota_id:
            errors.append({"row": i + 2, "error": "定额编号为空"})
            continue

        try:
            section = str(r.get("分部工程", "") or "") or guess_section(r)

            result = db.execute(
                text("""
                    INSERT INTO quotas (
                        quota_id, category, unit, quantity, project_name, work_content, section,
                        total_cost, labor_fee, material_fee, machinery_fee,
                        management_fee, tax, source_file
                    ) VALUES (
                        :quota_id, :category, :unit, :quantity, :project_name, :work_content, :section,
                        :total_cost, :labor_fee, :material_fee, :machinery_fee,
                        :management_fee, :tax, :source_file
                    )
                    ON CONFLICT (quota_id) DO UPDATE SET
                        category = EXCLUDED.category,
                        unit = EXCLUDED.unit,
                        quantity = EXCLUDED.quantity,
                        project_name = EXCLUDED.project_name,
                        work_content = EXCLUDED.work_content,
                        section = EXCLUDED.section,
                        total_cost = EXCLUDED.total_cost,
                        labor_fee = EXCLUDED.labor_fee,
                        material_fee = EXCLUDED.material_fee,
                        machinery_fee = EXCLUDED.machinery_fee,
                        management_fee = EXCLUDED.management_fee,
                        tax = EXCLUDED.tax,
                        source_file = EXCLUDED.source_file
                """),
                {
                    "quota_id": quota_id,
                    "category": r.get("类别") or r.get("category", ""),
                    "unit": r.get("计量单位") or r.get("单位", ""),
                    "quantity": r.get("数量") or "",
                    "project_name": r.get("项目名称") or "",
                    "work_content": r.get("工作内容") or "",
                    "section": section,
                    "total_cost": num(r.get("全费用")),
                    "labor_fee": num(r.get("人工费")),
                    "material_fee": num(r.get("材料费")),
                    "machinery_fee": num(r.get("机械费")),
                    "management_fee": num(r.get("管理费")),
                    "tax": num(r.get("增值税")),
                    "source_file": "excel_import",
                },
            )

            if result.rowcount == 0:
                skipped += 1
            else:
                success += 1
                # 统计分部
                top_section = section.split(" / ")[0] if section else "未分类"
                section_dist[top_section] += 1

            db.commit()
        except Exception as e:
            db.rollback()
            errors.append({"row": i + 2, "quota_id": quota_id, "error": str(e)})
            if len(errors) >= 10:
                break

    return {
        "total": total,
        "success": success,
        "skipped": skipped,
        "errors": errors[:10],
        "section_distribution": dict(section_dist),
    }


# ===== 路由 =====


@router.post("/import")
def import_quota_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: dict = Depends(require_role("admin")),
):
    """
    上传 Excel 定额文件并导入数据库。
    支持 .xlsx / .xls，自动识别表头并映射字段。
    返回导入报告。
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="文件名为空")

    suffix = file.filename.lower().split(".")[-1]
    if suffix not in ("xlsx", "xls"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    content = file.file.read()
    if len(content) > MAX_FILE_SIZE:
        return JSONResponse(
            status_code=413,
            content={"detail": f"文件大小超过限制（最大 10MB，当前 {len(content) // 1024 // 1024}MB）"},
        )

    try:
        records = parse_excel_to_records(content)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Excel 解析失败: {e}")

    if not records:
        raise HTTPException(status_code=422, detail="未找到有效数据行")

    # 导入
    result = import_records(db, records)

    # 生成报告
    report = {
        "imported_at": datetime.now().isoformat(),
        "filename": file.filename,
        "total_rows": result["total"],
        "success_count": result["success"],
        "skipped_count": result["skipped"],
        "error_count": len(result["errors"]),
        "errors": result["errors"],
        "section_distribution": result["section_distribution"],
    }

    _import_report["latest"] = report
    return report


@router.get("/report")
def get_import_report(
    db: Session = Depends(get_db),
    _: dict = Depends(require_role("admin")),
):
    """获取最新导入报告及数据库当前统计"""
    # 当前数据库统计
    quota_total = db.query(func.count(Quota.id)).scalar() or 0
    material_total = db.query(func.count(Material.id)).scalar() or 0

    # 分部分布
    section_rows = db.execute(
        text("""
            SELECT SPLIT_PART(section, ' / ', 1) as top_section, COUNT(*) as cnt
            FROM quotas WHERE section IS NOT NULL AND section != ''
            GROUP BY top_section ORDER BY cnt DESC
        """)
    ).fetchall()

    section_dist = {row.top_section: row.cnt for row in section_rows}

    return {
        "latest_import": _import_report.get("latest"),
        "db_stats": {
            "quota_count": quota_total,
            "material_count": material_total,
            "section_distribution": section_dist,
        },
    }
